
### DEAL WITH PANDAS ISSUES ###

import os
import pandas as pd
from openpyxl import load_workbook
from stpstone.cals.handling_dates import DatesBR


class DealingPd:
    '''
    DOCSTRING: COMMON FUNCTIONS TO DEAL WITH DATA IMPORT TO PANDAS DATAFRAMES OR TIMESERIES
    '''

    def append_df_to_Excel(self, filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False, bl_header=False, bl_index=0,
                           **to_Excel_kwargs):
        '''
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.
        Parameters:
        filename : File path or existing ExcelWriter
                    (Example: '/path/to/file.xlsx')
        df : dataframe to save to workbook
        sheet_name : Name of sheet which will contain DataFrame.
                    (default: 'Sheet1')
        startrow : upper left cell row to dump data frame.
                    Per default (startrow=None) calculate the last row
                    in the existing DF and write to the next row...
        truncate_sheet : truncate (remove and recreate) [sheet_name]
                        before writing DataFrame to Excel file
        to_Excel_kwargs : arguments which will be passed to `DataFrame.to_Excel()`
                            [can be dictionary]
        header : Y/N
        Returns: None
        '''
        # ignore [engine] parameter if it was passed
        if 'engine' in to_Excel_kwargs:
            to_Excel_kwargs.pop('engine')
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass
        if startrow is None:
            startrow = 0
        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow,
                    header=bl_header, index=bl_index)
        # save the workbook
        writer.save()

    def json_to_excel(self, json_path_name, xlsx_path_name):
        '''
        DOCSTRING: EXPORT JSON FILE TO EXCEL XLSX
        INPUTS: JSON AND XLSX COMPLETE NAME
        OUTPUTS: STATUS OF ACCOMPLISHMENT
        '''
        pd.read_json(json_path_name).to_excel(xlsx_path_name)
        if os.path.exists(xlsx_path_name):
            return 'OK'
        else:
            return 'NOK'

    def json_normalizer(self, json_path):
        '''
        DOCSTRING: JSON NORMALIZER FROM PANDAS
        INPUTS: JSON PATH
        OUTPUTS: JSON IN DATAFRAME
        '''
        return pd.read_json(json_path)

    def settingup_pandas(self):
        '''
        DOCSTRING: CONFIGURAÇÃO BÁSICA DO PANDAS
        INPUTS: COLUNA COMO STR
        OUTPUTS: -
        '''
        # Use 3 decimal places in output display
        pd.set_option("display.precision", 3)
        # Don't wrap repr(DataFrame) across additional lines
        pd.set_option("display.expand_frame_repr", False)
        # Set max rows displayed in output to 25
        pd.set_option("display.max_rows", 25)

    def dataframe_to_dict(self, dataframe_export, orientation='records'):
        '''
        REFERENCES: https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.to_dict.html
        DOCSTRING: EXPORT DATAFRAME TO JSON
        INPUTS: DATAFRAME TO BE EXPORTED, ORIENTATION(SPLIT, RECORDS (DEFAULT), INDEX, COLUMNS, VALUES, 
            TABLE)
        OUTPUTS: JSON
        '''
        return dataframe_export.to_dict(orient=orientation)

    def convert_datetime_columns(self, df_, list_col_date, bl_pandas_convertion=True):
        '''
        DOCSTRING: CONVERTING DATE COLUMNS TO DATETIME
        INPUTS: DATAFRAME, COLUMNS WITH DATE INFO
        OUTPUTS: DATAFRAME
        '''
        # checking wheter to covert through a pandas convertion, or resort to a excel format transformation
        #   of data in date column format
        if bl_pandas_convertion:
            for col_date in list_col_date:
                df_.loc[:, col_date] = pd.to_datetime(
                    df_[col_date], unit='s').dt.date
        else:
            # corventing list column dates to string type
            df_ = df_.astype(
                dict(zip(list_col_date, [str] * len(list_col_date))))
            # looping through each row
            for index, row in df_.iterrows():
                for col_date in list_col_date:
                    if '-' in row[col_date]:
                        ano = int(row[col_date].split(' ')[0].split('-')[0])
                        mes = int(row[col_date].split(' ')[0].split('-')[1])
                        dia = int(row[col_date].split(' ')[0].split('-')[2])
                        df_.loc[index, col_date] = DatesBR(
                        ).build_date(ano, mes, dia)
                    else:
                        df_.loc[index, col_date] = DatesBR(
                        ).excel_float_to_date(int(row[col_date]))
        # returning dataframe with date column to datetime forma
        return df_

    def merge_dfs_into_df(self, df_1, df_2, cols_list):
        '''
        DOCSTRING: MERGING TWO DFS INTO A NEW ONE, REMOVING THEIR INTERSECTION ON A LIST OF COLUMNS
        INPUTS: TWO DATAFRAMES AND A LIST WITH COLUMNS NAMES
        OUTPUTS: DATAFRAME
        '''
        intersec_df = pd.merge(df_1, df_2, how='inner', on=cols_list)
        merge_df = pd.merge(df_2, intersec_df, how='outer',
                            on=cols_list, indicator=True)
        merge_df = merge_df.loc[merge_df._merge == 'left_only']
        merge_df.dropna(how='all', axis=1, inplace=True)
        try:
            merge_df = merge_df.drop(columns='_merge')
        except:
            pass
        for column in merge_df.columns:
            if column in cols_list:
                pass
            else:
                merge_df = merge_df.rename(
                    columns={str(column): str(column)[:-2]})
        return merge_df
